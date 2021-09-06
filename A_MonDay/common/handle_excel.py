import openpyxl
import unittest


class Handle_excel():
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    # 读取Excel中的数据
    def read_excel(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)

        title = [i.value for i in res[0]]
        cases = []

        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    # 回写数据到Excel中
    def write_excel(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
