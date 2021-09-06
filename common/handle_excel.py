import unittest
import openpyxl


class Handle_excel():

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)

        cases = []
        title = [i.value for i in res[0]]

        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
